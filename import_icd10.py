"""
Bulk-loads ICD-10-CM codes from a CMS "Section 111 Valid ICD-10 codes"
spreadsheet into the DiagnosisCode catalog.

Codes in the source file come without the decimal point (e.g. "A000").
This script re-inserts it in the standard ICD-10-CM display format
("A00.0") so it lines up with how clinicians actually read/write codes.

Usage:
    python import_icd10.py icd10_source.xlsx
    python import_icd10.py icd10_source.xlsx --sheet "Valid ICD10 FY2026 & NF Exclude"
    python import_icd10.py icd10_source.xlsx --replace   # wipe existing catalog first

Re-run this whenever CMS publishes an updated file (usually each October).
"""
import argparse

from openpyxl import load_workbook

from app import create_app
from app.extensions import db
from app.models import DiagnosisCode

DEFAULT_SHEET = "Valid ICD10 FY2026 & NF Exclude"
BATCH_SIZE = 5000


def format_code(raw):
    raw = str(raw).strip().upper()
    if len(raw) > 3 and "." not in raw:
        return f"{raw[:3]}.{raw[3:]}"
    return raw


def parse_rows(xlsx_path, sheet_name):
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    rows = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        code_raw = row[0]
        short_desc = row[1] if len(row) > 1 else None
        long_desc = row[2] if len(row) > 2 else None

        code = format_code(code_raw)
        if code in seen:
            continue
        seen.add(code)

        description = (short_desc or long_desc or "").strip()
        if not description:
            continue

        rows.append({
            "code": code,
            "description": description[:255],
            "category": None,
            "is_active": True,
        })
    return rows


def run(xlsx_path, sheet_name, replace_existing):
    app = create_app()
    with app.app_context():
        print(f"Reading {xlsx_path} (sheet: {sheet_name})...")
        rows = parse_rows(xlsx_path, sheet_name)
        print(f"Parsed {len(rows)} unique ICD-10 codes.")

        if replace_existing:
            deleted = DiagnosisCode.query.delete()
            db.session.commit()
            print(f"Cleared {deleted} existing diagnosis codes.")

        existing_codes = {c[0] for c in db.session.query(DiagnosisCode.code).all()}
        new_rows = [r for r in rows if r["code"] not in existing_codes]
        print(f"{len(new_rows)} codes are new; inserting in batches of {BATCH_SIZE}...")

        for i in range(0, len(new_rows), BATCH_SIZE):
            batch = new_rows[i:i + BATCH_SIZE]
            db.session.bulk_insert_mappings(DiagnosisCode, batch)
            db.session.commit()
            print(f"  inserted {min(i + BATCH_SIZE, len(new_rows))}/{len(new_rows)}")

        total = DiagnosisCode.query.count()
        print(f"Done. Diagnosis catalog now has {total} codes.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import ICD-10-CM codes into the diagnosis catalog.")
    parser.add_argument("xlsx_path", help="Path to the CMS Section 111 valid ICD-10 xlsx file")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Sheet name to read (defaults to the FY valid-codes sheet)")
    parser.add_argument("--replace", action="store_true", help="Delete all existing diagnosis codes before importing")
    args = parser.parse_args()
    run(args.xlsx_path, args.sheet, args.replace)
